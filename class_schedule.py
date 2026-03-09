import tkinter as tk
from tkinter import filedialog
import openpyxl
import warnings
import sys
import uuid
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from icalendar import Calendar, Event, vDatetime, vRecur

# Configuration constants for Workday Excel export format
MAX = 100  # Maximum rows/columns to search. Increase if spreadsheet is truncated.
COL_SKIP = 1  # Workday exports: skip 1 label column before headers start
ROW_SKIP = 2  # Workday exports: skip 2 rows (title + blank) before headers start

WEEKDAYS = {
    'M': "MO",
    'T': "TU",
    'W': "WE",
    'R': "TH",
    'F': "FR"
}

# Required column headers - script will fail without these
REQUIRED_HEADERS = [
    'Course Listing',
    'Section',
    'Meeting Patterns',
    'Start Date',
    'End Date'
]

# Optional headers - used for descriptions but not critical
OPTIONAL_HEADERS = [
    'Instructor',
    'Delivery Mode',
    'Instructional Format'
]


def get_filename():
    """
    Open a system dialog to allow the user to select a .xlsx file.

    Returns:
        str: Path to selected file, or empty string if cancelled
    """
    print("\nPlease select an excel file to read.")

    # start up tkinter for the file dialog
    root = tk.Tk()
    root.withdraw()

    # acquire the input file with OS file dialog
    fname = filedialog.askopenfilename(filetypes=[('Microsoft Excel Files', '*.xlsx')])

    # close tkinter now that we're done with it
    root.destroy()
    return fname


def parse_spreadsheet(fname, expected_headers=None):
    """
    Parse .xlsx file and extract section information with validation.

    Each section is assigned a unique UUID and parsed according to Workday
    export format. If expected_headers is provided, validates that this file's
    headers match (for multi-file consistency).

    Args:
        fname (str): Path to the .xlsx file
        expected_headers (list, optional): Expected column headers from previous file

    Returns:
        tuple: (sections, headers) where:
            - sections (list): List of section dictionaries, each with UUID
            - headers (list): Column headers found in this file
            - Returns (None, None) if headers don't match expected_headers
    """
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
    # try to open the file
    try:
        print("Loading excel spreadsheet...")
        wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
        ws = wb.active
    except openpyxl.utils.exceptions.InvalidFileException:
        # exits if dialog was cancelled
        print("Invalid file!")
        sys.exit()
    
    start_row = ROW_SKIP + 1
    start_col = COL_SKIP + 1

    # Validate that skip values are correct
    first_header = ws.cell(start_row, start_col).value
    if first_header is None:
        print("\nError: First header cell is empty!")
        print(f"Expected headers at row {start_row}, column {start_col}")
        print(f"Current settings: ROW_SKIP={ROW_SKIP}, COL_SKIP={COL_SKIP}")
        print("\nPlease check:")
        print("  1. Your Excel file matches the expected Workday export format")
        print("  2. Adjust ROW_SKIP/COL_SKIP constants if needed (see README.md)")
        sys.exit(1)

    print("Finding headers...")
    # parses the spreadsheet to find column names
    headers = []
    for col in range(MAX):
        # traverses the third row, starting on the second column
        cell = ws.cell(start_row, col + start_col).value
        # only adds the cell if the column exists
        if cell is not None:
            headers.append(cell)

    # Validate that all required headers are present
    missing_headers = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing_headers:
        print("\nError: Missing required headers in spreadsheet!")
        print(f"Missing: {', '.join(missing_headers)}")
        print(f"\nFound headers: {', '.join(headers)}")
        print("\nPlease check:")
        print("  1. Your Excel file is a Workday export (Academics → View My Courses → Export to Excel)")
        print("  2. The ROW_SKIP/COL_SKIP constants are correctly configured")
        sys.exit(1)

    # If expected_headers provided, validate they match
    if expected_headers is not None:
        if set(headers) != set(expected_headers):
            print("\nError: Headers don't match previous file!")
            print(f"Expected: {', '.join(expected_headers)}")
            print(f"Found:    {', '.join(headers)}")
            print("\nAll files must have the same column structure.")
            print("This file will be skipped.")
            return None, None

    start_row += 1  # moves down to the next row for section data

    print("Finding sections...")
    # parses section info from each row
    sections = []
    rows_read = 0
    for row in range(MAX):

        if ws.cell(row + start_row, start_col).value is None:
            break

        section = {'UUID': uuid.uuid4().hex.upper()}
        for col in range(len(headers)):
            section[headers[col]] = ws.cell(row + start_row, col + start_col).value

        if section['Meeting Patterns'] is not None:
            fields = section['Meeting Patterns'].split(" | ")

            time_str = fields[1].split(" - ")
            section['Start Time'] = datetime.strptime(time_str[0], r"%H:%M")
            section['End Time'] = datetime.strptime(time_str[1], r"%H:%M")

            section['Location'] = fields[2]
            section['Meeting Patterns'] = fields[0].split("-")

        sections.append(section)
        rows_read = row + 1

    print(f"Read {rows_read} rows from spreadsheet")

    # Warn if we hit the MAX limit
    if rows_read >= MAX:
        print(f"Warning: Reached MAX limit of {MAX} rows - spreadsheet may be truncated!")
        print(f"To process more rows, increase MAX constant (currently {MAX})")

    return sections, headers


def group_data(sections):
    """
    Group sections into courses, and courses into time frames.

    Sections with the same Course Listing are grouped together. Courses
    with the same start/end dates are grouped into time frames.

    Args:
        sections (list): List of section dictionaries

    Returns:
        tuple: (courses, time_frames) where:
            - courses (dict): Maps course name to list of sections
            - time_frames (dict): Maps (start_date, end_date) tuple to list of courses
    """
    print("Grouping sections into courses...")
    # groups sections into courses
    courses = {}
    for section in sections:
        course = section['Course Listing']  # identifies the course name
        if course in courses.keys():
            courses[course].append(section)  # appends to entry if not new
        else:
            courses[course] = [section]  # makes new entry if new

    print("Grouping courses into time frames...")
    # groups courses into time frames
    time_frames = {}
    for course, course_sections in courses.items():
        # a time frame is tuple of start and end date
        time_frame = (course_sections[0]['Start Date'], course_sections[0]['End Date'])
        if time_frame in time_frames:  # appends to entry if not new
            time_frames[time_frame].append(course)
        else:  # makes new entry if new
            time_frames[time_frame] = [course]
    
    return courses, time_frames


def verify_scheduling(sections, courses, time_frames):
    """
    Verify sections have schedules and remove unscheduled items.

    Removes sections without Meeting Patterns, courses with no scheduled
    sections, and time frames with no courses. Modifies inputs in place.

    Args:
        sections (list): List of section dictionaries (modified in place)
        courses (dict): Course name to sections mapping (modified in place)
        time_frames (dict): Time frame to courses mapping (modified in place)
    """
    print("\nVerifying schedule data...")
    # for each time frame (iterating over key list copy)
    for time_frame in list(time_frames.keys()):
        # get the list of courses in the time frame
        tf_courses = time_frames[time_frame]

        # iterate over a copy of the course list
        for course in list(tf_courses):
            # get the list of sections in the course
            course_sections = courses[course]

            # iterate over a copy of the sections list
            for section in list(course_sections):

                # check whether the section has a schedule
                if section['Meeting Patterns'] is None:
                    # display a message
                    code = " ".join(section['Section'].split()[:2])
                    print(f"Discarded section {code} because not scheduled.")
                    course_sections.remove(section)  # remove it from the course
                    sections.remove(section)  # remove it from the main list

            # check whether the course has any section left
            if len(course_sections) == 0:
                # display a message
                code = " ".join(course.split()[:2])
                print(f"Discarded course {code} due to no scheduled sections.")
                tf_courses.remove(course)  # remove it from the time frame
                del courses[course]  # remove it from the main dict

        # check whether the time frame has any courses left
        if len(tf_courses) == 0:
            # display a message
            time_str = f"{datetime.strftime(time_frame[0], r"%Y-%m-%d")} to {datetime.strftime(time_frame[1], r"%Y-%m-%d")}"
            print(f"Discarded time frame from {time_str} due to no remaining courses.")
            del time_frames[time_frame]  # remove from the dict


def print_data_summary(sections, courses, time_frames):
    """
    Print summary statistics of loaded data.

    Args:
        sections (list): List of section dictionaries
        courses (dict): Course name to sections mapping
        time_frames (dict): Time frame to courses mapping
    """
    # displays information about the courses
    print(f"\n{len(sections)} sections, {len(courses.keys())} courses, and {len(time_frames.keys())} time frames found!")


def print_tree_view(courses, time_frames):
    """
    Print hierarchical tree view of time frames, courses, and sections.

    Args:
        courses (dict): Course name to sections mapping
        time_frames (dict): Time frame to courses mapping
    """

    print()
    print("=" * 60)
    print("Tree view of time frames, courses, and sections")
    print("=" * 60)

    # for each time frame
    for time_frame, tf_courses in time_frames.items():
        # print the time frame
        print(f"\n{datetime.strftime(time_frame[0], r"%Y-%m-%d")} to {datetime.strftime(time_frame[1], r"%Y-%m-%d")}:")

        # for each course in the time frame
        for i, course in enumerate(tf_courses):
            # print the course name and gutter tree
            print("│")
            last = i == len(tf_courses) - 1
            char = '└' if last else '├'
            print(f"{char}── {course}")

            # identify the sections
            course_sections = courses[course]

            # for each section in the course
            for i, section in enumerate(course_sections):
                # print section name and gutter tree
                char = '└' if i == len(course_sections) - 1 else '├'
                tree = ' ' if last else '│'
                code = " ".join(section['Section'].split()[:2])
                print(f"{tree}   {char}── {code:<14}", end="")

                # print meeting patterns
                out = ""
                if section['Meeting Patterns'] is not None:
                    out = ", ".join(section['Meeting Patterns'])
                print(f"    {out:<13}", end="")

                # print start time
                out = "--:--"
                if 'Start Time' in section.keys():
                    out = section['Start Time'].strftime(r"%H:%M")
                print(f"    {out}", end="")

                # print end time
                out = "--:--"
                if 'End Time' in section.keys():
                    out = section['End Time'].strftime(r"%H:%M")
                print(f" - {out}")


def select_sections(time_frames, courses):
    """
    Interactively prompt user to select sections for calendar export.

    Allows drill-down selection: time frame → course → individual section.
    User can approve entire time frames, individual courses, or specific sections.

    Args:
        time_frames (dict): Time frame to courses mapping
        courses (dict): Course name to sections mapping

    Returns:
        list: Approved section dictionaries selected by user
    """
    approved_sections = []

    print()
    print("=" * 60)
    print("Select sections to export to calendar")
    print("=" * 60)

    # iterate over each time frame
    for time_frame, tf_courses in time_frames.items():
        time_str = f"{datetime.strftime(time_frame[0], r'%Y-%m-%d')} to {datetime.strftime(time_frame[1], r'%Y-%m-%d')}"
        print(f"\nTime frame: {time_str}")

        while True:
            choice = input("  (y)es / (n)o / (s)pecific courses? ").lower().strip()
            if choice in ['y', 'n', 's']:
                break
            print("Invalid input. Please enter y, n, or s.")

        if choice == 'y':
            # approve all sections in this time frame
            for course in tf_courses:
                approved_sections.extend(courses[course])
            print(f"  ✓ Added all courses in this time frame")

        elif choice == 's':
            # drill down to course selection
            for course in tf_courses:
                print(f"\n    Course: {course}")

                while True:
                    choice = input("      (y)es / (n)o / (s)pecific courses? ").lower().strip()
                    if choice in ['y', 'n', 's']:
                        break
                    print("    Invalid input. Please enter y, n, or s.")

                if choice == 'y':
                    # approve all sections in this course
                    approved_sections.extend(courses[course])
                    print(f"      ✓ Added all sections in this course")

                elif choice == 's':
                    # drill down to section selection
                    course_sections = courses[course]
                    for section in course_sections:
                        code = " ".join(section['Section'].split()[:2])

                        # show schedule info
                        days = ", ".join(section['Meeting Patterns'])
                        start = section['Start Time'].strftime(r"%H:%M")
                        end = section['End Time'].strftime(r"%H:%M")

                        print(f"\n        Section: {code} ({days} {start}-{end})")

                        while True:
                            choice = input("          (y)es / (n)o? ").lower().strip()
                            if choice in ['y', 'n']:
                                break
                            print("        Invalid input. Please enter y or n.")

                        if choice == 'y':
                            approved_sections.append(section)
                            print(f"          ✓ Added section")
                        else:
                            print(f"          ✗ Skipped section")

                else:
                    # choice == 'n' means skip this course
                    print(f"      ✗ Skipped course")

        else:
            # choice == 'n' means skip this time frame
            print(f"  ✗ Skipped time frame")
    
    return approved_sections


def generate_calendar(sections):
    """
    Generate iCalendar object from approved sections.

    Each section becomes a calendar event with its UUID as the event UID.
    Uses defensive .get() for optional fields (Instructor, Delivery Mode, etc.).

    Args:
        sections (list): List of approved section dictionaries

    Returns:
        Calendar: iCalendar object ready for export
    """
    print("\nGenerating iCalendar data...")
    cal = Calendar()
    cal.add('prodid', '-//Class Schedule Importer//mxm.dk//')
    cal.add('version', '2.0')

    for section in sections:
        event = Event()
        event.add('uid', section['UUID'])
        event.add('summary', section.get('Section', 'Unknown Section'))
        event.add('location', section.get('Location', 'TBD'))

        tz = ZoneInfo('America/New_York')
        dtstart = datetime.combine(section['Start Date'].date(), section['Start Time'].time(), tzinfo=tz).astimezone(timezone.utc)
        dtend = datetime.combine(section['Start Date'].date(), section['End Time'].time(), tzinfo=tz).astimezone(timezone.utc)
        event.add('dtstart', dtstart)
        event.add('dtend', dtend)
        event.add('dtstamp', datetime.now(timezone.utc))

        # Build description with optional fields that read naturally when missing
        delivery_mode = section.get('Delivery Mode', 'Unknown mode')
        instructional_format = section.get('Instructional Format', 'unknown format').lower()
        instructor = section.get('Instructor', 'unknown professor')
        description = f"{delivery_mode}, {instructional_format} with {instructor}.\n\nGenerated by Class Schedule Importer."
        event.add('description', description)

        cal.add_component(event)

    return cal


def save_calendar(cal):
    """
    Save iCalendar object to .ics file via file dialog.

    Args:
        cal (Calendar): iCalendar object to save
    """
    print("Please select a location to save the iCalendar file.")
    # start up tkinter for the file dialog
    root = tk.Tk()
    root.withdraw()

    # acquire the output file with OS file dialog
    fname = f"class_schedule_{datetime.now().strftime(r'%Y%m%d')}.ics"
    fname = filedialog.asksaveasfilename(defaultextension=".ics", initialfile=fname, filetypes=[('iCalendar Files', '*.ics')])

    root.destroy()

    # exits if dialog was cancelled
    if fname == "":
        print("Save cancelled.")
        return
    # ensure the filename ends with .ics
    if not fname.endswith(".ics"):
        fname += ".ics"
    
    # write the calendar to the file
    with open(fname, 'wb') as f:
        f.write(cal.to_ical())
    print(f"iCalendar data saved to {fname}")


def main():
    print("Welcome to Class Schedule Importer!")
    print("For usage instructions, see README.md")

    # Load files in a loop
    all_sections = []
    headers = None

    while True:
        fname = get_filename()
        if not fname:  # User cancelled
            break

        result = parse_spreadsheet(fname, headers)
        if result == (None, None):  # Header mismatch
            continue

        sections, headers = result
        all_sections.extend(sections)
        print(f"Total sections loaded: {len(all_sections)}")

        if input("\nLoad another file? (y/n): ").lower().strip() != 'y':
            break

    if not all_sections:
        print("No sections loaded. Exiting.")
        sys.exit()

    courses, time_frames = group_data(all_sections)
    print_data_summary(all_sections, courses, time_frames)
    verify_scheduling(all_sections, courses, time_frames)
    print_data_summary(all_sections, courses, time_frames)
    print_tree_view(courses, time_frames)
    approved_sections = select_sections(time_frames, courses)
    print(f"\n{len(approved_sections)} sections out of {len(all_sections)} approved for export.")
    calendar = generate_calendar(approved_sections)
    save_calendar(calendar)
    print("\nThank you for using Class Schedule Importer!")


if __name__ == "__main__":
    main()
